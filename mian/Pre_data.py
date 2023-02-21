# author: code_king
# time: 2022/2/19 12:37
# file: Pre_data.py
"""
还是5%异常来处理
前面阶段的输入和输出是否都是下一个阶段的输入’和‘前面阶段的输入是下一个阶段的输入’两种方式都尝试一下
"""
import copy
import os
import warnings

import joblib
import numpy as np
import pandas as pd
from numpy import unique
from sklearn.metrics import accuracy_score

from common_utils.DimensionalityReduction import DataDelete
from common_utils.plot_error_distribution import write_images
from model_utils.NewPreDeal_Tools import Del_deletion_data, get_final_useablecols, all_ydata, \
    delAndGetCols, get_former_Ydata, convert_to_num, deal_verify_ydata, Deal_sorted_Ydata
from plot_utils.build_images.init_build_images import gen_images

# warnings.filterwarnings(action='always', category=UserWarning)
warnings.filterwarnings(action='ignore')
# base_data_dirs = "../datas"
# 这个pandas处理数据效果不太好 建议用numpy
init_data = pd.read_excel(io=f'data_original.xlsx', sheet_name='所有数据')
# Preddata = pd.read_excel(io=r'data_original.xlsx', sheet_name='8000D数据标准化-预测用')
# 用来查看数据结果的excel对象
from openpyxl import load_workbook

filename = 'data_original.xlsx'
data = load_workbook(f'{filename}')
sheetnames = data.sheetnames
# create_sheet
if len(sheetnames) < 4:
    sheetnames.append('预测结果')
    data.create_sheet(sheetnames[2], len(sheetnames))
    # 赋值sheet
    # sheet=data[sheetnames[0]]
    # content=data.copy_worksheet(sheet)
    data.save(f'{filename}')
# 预测结果的excel表
table = data[sheetnames[-1]]
sheetnames = data.sheetnames
table = data[sheetnames[-1]]

# 获取数据表对象(建模和与预测的数据)
init_data = np.array(init_data)


# Preddata = np.array(Preddata)


# 写入excel文件
def write_to_excel(start_index, table, current_column, content):
    # 标签
    for i in range(len(content)):
        table.cell(start_index, current_column).value = content[i]
        start_index += 1


def save_model(model, model_name, X_train, X_test, y_train, y_test, train_numbers):
    """
    :param model: 需要训练的模型
    :param X_train:
    :param X_test:
    :param y_train:
    :param y_test:
    :return:
    """
    # 训练次数
    if not os.path.exists(f"models/{model_name}"):
        os.makedirs(name=f"models/{model_name}", exist_ok=True)
    acc_dic = {}
    for i in range(11):
        model.fit(X_train, y_train.astype(float))
        # 训练集准确率
        model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
        # print('Rfc_train_acc准确率：', model_train_acc)
        model_pred = model.predict(X_test)
        model_acc = accuracy_score(y_test.astype(float), model_pred)
        # 取最小的
        model_train_acc = min(model_train_acc, model_acc)
        acc_dic.update({f"model[{i}]": model_train_acc})
        # 模型保存
        joblib.dump(model, f'models/{model_name}/model[{i}].pkl')
    for i in range(11, train_numbers):
        # temp_keys 是键的按照值从小到打排序的列表 [model[1],model[5],...]
        temp_keys = sorted(acc_dic, reverse=True)
        model.fit(X_train, y_train.astype(float))
        # 训练集准确率
        model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
        # print('Rfc_train_acc准确率：', model_train_acc)
        model_pred = model.predict(X_test)
        model_acc = accuracy_score(y_test.astype(float), model_pred)
        model_train_acc = min(model_train_acc, model_acc)
        # 替换一个最小的模型
        for key_item_index in range(0, len(temp_keys)):
            if key_item_index != len(temp_keys) - 1:
                if acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc < acc_dic[
                    f"{temp_keys[key_item_index + 1]}"]:
                    acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
                    joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')
            elif acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc:
                acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
                joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')

    # model1 = joblib.load(filename="filename.pkl")


def get_pred_data(model_name, X_verify_data):
    """
    :param model_name: 模型名字
    :return: pred_data
    """
    # 预测数据
    all_models = os.listdir(f"models/{model_name}")
    model_pred_list = []
    for i in all_models:
        cursor_model = joblib.load(filename=f"models/{model_name}/{i}")
        model_pred_list.append(cursor_model.predict(X_verify_data))
    model_pred_list = np.array(model_pred_list)
    temp_pred_list = []
    # 循环遍历列
    for pred_list_index in range(0, model_pred_list.shape[1]):
        # 预测结果统一,那个数多就等于哪个
        zero_count = np.where(model_pred_list[:, pred_list_index] == 0)[0].shape[0]
        negative_count = np.where(model_pred_list[:, pred_list_index] == -1)[0].shape[0]
        positive_count = np.where(model_pred_list[:, pred_list_index] == 1)[0].shape[0]
        final_dic = {}
        final_dic.update({"0": zero_count})
        final_dic.update({"-1": negative_count})
        final_dic.update({"1": positive_count})
        # 排序，第一个数预测的最多
        final_result = -1
        if final_dic[f"{final_result}"] < final_dic["0"]:
            final_result = 0
        if final_dic[f"{final_result}"] < final_dic["1"]:
            final_result = 1
        temp_pred_list.append(final_result)
    pred_data = np.array(temp_pred_list)
    return pred_data


def get_train_test_acc(model, X_train, y_train, X_test, y_test):
    """
    :param model:
    :param X_train:
    :param y_train:
    :param X_test:
    :param y_test:
    :return: model_train_acc, model_acc
    """
    model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
    model_pred = model.predict(X_test)
    model_acc = accuracy_score(y_test.astype(float), model_pred)
    return model_train_acc, model_acc


def get_develop_pred_data(excel_data=None, boundary_x=None, boundary_y=None):
    """
    :param excel_data:
    :param boundary_x:
    :param boundary_y:
    :return:
    """
    return excel_data[boundary_x[0]:boundary_x[-1], boundary_y[0]:boundary_y[-1]]


def get_train_test_data(x_data=None, y_data=None):
    """ 按照奇偶分
    :param X_data:
    :param Y_data:
    :return: X_train, X_test, y_train, y_test
    """
    # 转化成ndarray操作
    X_train = []
    X_test = []
    Y_train = []
    Y_test = []
    for i in range(x_data.shape[0]):
        if i % 2 == 0:
            X_train.append(x_data[i, :])
            Y_train.append(y_data[i])
        else:
            X_test.append(x_data[i, :])
            Y_test.append(y_data[i])
    X_train = np.array(X_train)
    X_test = np.array(X_test)
    Y_train = np.array(Y_train)
    Y_test = np.array(Y_test)
    return X_train, X_test, Y_train, Y_test


def get_usable_data(x_data=None, y_data=None):
    """ 获取删除无效行的数据
    :param x_data: 删除无效列的
    :param y_data: 删除无效列的
    :return: 删除无效行以后的数据
    """
    # 先按照列合并
    temp_OriginalXdata = np.column_stack((x_data, y_data))
    # 删除缺失的行
    temp_OriginalXdata = Del_deletion_data(temp_OriginalXdata, 0)
    # X_data,Y_data 是处理好(行列均可用)的x,y
    X_data = temp_OriginalXdata[:, :-1]
    Y_data = temp_OriginalXdata[:, -1]
    print(
        f'空列表就是删除干净了.**{list(set(np.where(np.isnan(temp_OriginalXdata.astype(float)) == True)[0].tolist()))}', )
    return X_data, Y_data


def get_invert_y(y_develop_data=None, y_verify_data=None):
    """
    :param y_develop_data:建模的Y(源数据，未转化-1 0 1)
    :param y_verify_data: 预测的Y(源数据，未转化-1 0 1)
    :return: 转化好的 Y_develop_data,Verify_Ydata
    """
    # 获取 建模和预测转化好的Y
    Y_develop_data_copy = copy.deepcopy(y_develop_data)
    y_develop_data, Y_data_boundsMin, Y_data_boundsMax = Deal_sorted_Ydata(data=Y_develop_data_copy)
    # 获取预测数据的Y_data Verify_Ydata = Pred_Y_data
    # 保留原始的数据
    Y_verify_data_copy = copy.deepcopy(y_verify_data)
    # 验证（预测）的Y
    Verify_Ydata = deal_verify_ydata(Y_verify_data_copy, Y_data_boundsMin, Y_data_boundsMax)
    # 查看是否替换成功
    print("查看是否替换成功：", [unique(y_develop_data), Y_data_boundsMin, Y_data_boundsMax],
          np.unique(Verify_Ydata))
    return y_develop_data, Verify_Ydata


def develop_model(x_develop_data=None, x_pred_data=None, y_develop_data=None, y_pred_data=None, final_cols=None,
                  y_filename=None, traverse_index=None):
    """
    :param x_develop_data:
    :param x_pred_data:
    :param y_develop_data:
    :param y_pred_data:
    :param final_cols: 可用的列
    :param y_filename: 根据月份的y的文件位置
    :param traverse_index: 遍历次数，可用乘一个固定数，这样让excel的数据加进去
    :return:
    """
    the_column = 88
    for index in range(len(final_cols)):
        current_index = final_cols[index]
        y_develop_data_curr = y_develop_data[current_index]
        y_verify_data_curr = y_pred_data[current_index]
        # 拷贝原数据，这个画图还是需要使用以前的数据
        Y_develop_data_orinal = copy.deepcopy(y_develop_data_curr)
        Y_verify_data_orinal = copy.deepcopy(y_verify_data_curr)
        # 已经处理过列了，现在获取行可用的 建模数据
        X_develop_data, Y_develop_data = get_usable_data(x_data=x_develop_data, y_data=y_develop_data_curr)
        # 处理验证（预测）的X_data,Y_data
        X_verify_data, Y_verify_data = get_usable_data(x_data=x_pred_data, y_data=y_verify_data_curr)
        Y_develop_data, Verify_Ydata = get_invert_y(y_develop_data=Y_develop_data, y_verify_data=Y_verify_data)
        # 手动划分数据集，各自一半，按照奇偶
        X_train, X_test, y_train, y_test = get_train_test_data(x_data=X_develop_data, y_data=Y_develop_data)
        # 验证集
        X_verify_data = X_verify_data
        Y_verify_data = Verify_Ydata
        # 开始训练模型
        from sklearn.ensemble import RandomForestClassifier
        # 设置训练次数
        train_numbers = 100
        # 获取模型  随机森林
        Rfc = RandomForestClassifier()
        model_name = "Rfc"
        save_model(Rfc, model_name, X_train, X_test, y_train, y_test, train_numbers)
        # 随便取一个作为训练集的结果
        current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
        Rfc_train_acc, Rfc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)
        Rfc_pred = get_pred_data("Rfc", X_verify_data)
        verify_y1 = accuracy_score(Y_verify_data.astype(float), Rfc_pred)

        # todo 每个阶段不一样，后面这个会变动
        # gen_images(iter_left_data=base_Xdata, iter_right_verify=Verify_Xdata, files_name="x_images", titles=all_titles)
        write_images(data_left=Y_develop_data_orinal, data_right=Y_verify_data_orinal,
                     names=f"{y_filename}/{all_titles[the_column + current_index - 1]}.png", score=verify_y1)
        # 计算准确率
        print('验证集Rfc_acc准确率:', verify_y1)
        #  标签 都加20
        write_to_excel(start_index=5 + traverse_index * 10, table=table, current_column=1,
                       content=[f"({y_filename})训练集个数", "随机森林（Rfc）准确率",
                                "测试集个数", "随机森林（Rfc）准确率",
                                "验证集个数", "随机森林（Rfc）准确率", ])
        # 将训练集写入到excel中
        write_to_excel(start_index=5 + traverse_index * 10, table=table, current_column=the_column + index,
                       content=[X_train.shape[0], Rfc_train_acc,
                                X_test.shape[0], Rfc_acc,
                                X_verify_data.shape[0], verify_y1])
        data.save(filename)


def write_x_images(develop_x_title_data=None, develop_x_data=None, verify_x_data=None, x_filename="x_images"):
    """
    :param develop_x_title_data: 带标题的
    :param develop_x_data:
    :return:
    """
    # 将 建模的X 和 预测的X 传入即可，可先将标题带上
    title_data = list(develop_x_title_data)
    for i in range(len(develop_x_title_data)):
        write_images(data_left=develop_x_data[:, i], data_right=verify_x_data[:, i],
                     names=f"{x_filename}/{title_data[i]}.png")


if __name__ == '__main__':
    # 将表的内容 7个数据表 挨个处理
    # 获取源建模的X数据 传入的对象是获取的excel对象
    # 获取标题内容
    all_titles = init_data[1, :]
    process_title = init_data[1, 3:]
    # 根据月份边际去遍历
    x_boundary_right = [320, 343, 370, 393, 421, 443, 455]
    x_boundary_left = [320, 343, 370, 393, 421, 443, 455]
    index = 0
    for boundary_index in range(len(x_boundary_right) - 1):
        # 建模数据
        develop_x_data = get_develop_pred_data(excel_data=init_data, boundary_x=[3, x_boundary_right[boundary_index]],
                                               boundary_y=[4, 87])
        develop_x_title_data = get_develop_pred_data(excel_data=init_data,
                                                     boundary_x=[1, x_boundary_right[boundary_index]],
                                                     boundary_y=[4, 87])
        develop_y_data = get_develop_pred_data(excel_data=init_data, boundary_x=[3, x_boundary_right[boundary_index]],
                                               boundary_y=[87, 106])

        # 预测数据
        pred_x_data = get_develop_pred_data(excel_data=init_data, boundary_x=[x_boundary_right[boundary_index],
                                                                              x_boundary_right[boundary_index + 1]],
                                            boundary_y=[4, 87])
        pred_y_data = get_develop_pred_data(excel_data=init_data, boundary_x=[x_boundary_right[boundary_index],
                                                                              x_boundary_right[boundary_index + 1]],
                                            boundary_y=[87, 106])

        # 统一关键字
        the_Former_data, the_Verify_TABLE_data, del_list = convert_to_num(develop_x_data, pred_x_data)
        develop_x_title_data = np.delete(develop_x_title_data, del_list, axis=1)
        # 删除缺失值过多的列，并保存del_cols
        base_Xdata, del_cols = delAndGetCols(the_Former_data)
        # 用del_cols删除验证集X不需要的列
        Verify_Xdata = np.delete(the_Verify_TABLE_data, del_cols, axis=1)
        develop_x_title_data = np.delete(develop_x_title_data, del_cols, axis=1)
        # 降维处理,这个地方只从训练集判断相关性，然后统一降维·
        use_able_x_cols = DataDelete(base_Xdata, 0.80)
        # 所用可用列的X： base_Xdata
        base_Xdata = base_Xdata[:, use_able_x_cols]
        # 获取实际可用列的（行会和Y的数据一起删除） 预测的X数据 Verify_Xdata
        Verify_Xdata = Verify_Xdata[:, use_able_x_cols]
        # 获取可用的列
        develop_x_title_data = develop_x_title_data[0, use_able_x_cols]
        # 处理预测
        final_cols = get_final_useablecols(develop_y_data, pred_y_data)
        # 获取可以用“列”的Y_data
        Original_YdataList, Pred_YdataList = all_ydata(final_cols, develop_y_data, pred_y_data)
        # gen_images(iter_left_data=base_Xdata, iter_right_verify=Verify_Xdata, files_name="x_images", titles=all_titles)
        write_x_images(develop_x_title_data=develop_x_title_data, develop_x_data=base_Xdata, verify_x_data=Verify_Xdata,
                       x_filename=f"x_images[{x_boundary_right[boundary_index]}]")
        # 依次遍历每一个Y
        develop_model(x_develop_data=base_Xdata, x_pred_data=Verify_Xdata, y_develop_data=Original_YdataList,
                      y_pred_data=Pred_YdataList, final_cols=final_cols,
                      y_filename=f"y_images[{x_boundary_right[boundary_index]}]", traverse_index=index)
        index += 1
    print("运行结束！")

    # for i in range(iter_length):
    #     # 统一关键字
    #     the_Former_data, the_Verify_TABLE_data = convert_to_num(all_former_data[i], all_verify_data[i])
    #     # 删除缺失值过多的列，并保存del_cols
    #     # tempFormerTwo_data = delAndGetCols(FormerTwo_data)
    #     base_Xdata, del_cols = delAndGetCols(the_Former_data)
    #     # 用del_cols删除验证集X不需要的列
    #     Verify_Xdata = np.delete(the_Verify_TABLE_data, del_cols, axis=1)
    #     # 降维处理,这个地方只从训练集判断相关性，然后统一降维·
    #     use_able_x_cols = DataDelete(base_Xdata, 0.80)
    #     base_Xdata = base_Xdata[:, use_able_x_cols]
    #     Verify_Xdata = Verify_Xdata[:, use_able_x_cols]
    #     # 合并数据一起作为训练集
    #     # all_train_x_data = np.r_[base_Xdata, Verify_Xdata]
    #
    #     # base_Xdata,del_cols
    #     # %%
    #     # 获取源建模的数据Y_data
    #     Original_TableTwoYdata, Original_TableThreeYdata, Original_TableFiveYdata, Original_TableSixYdata, Original_TableSevenYdata = get_former_Ydata(
    #         init_data)
    #     # 获取预测数据的Y_data
    #     Predict_TableTwoYdata, Predict_TableThreeYdata, Predict_TableFiveYdata, Predict_TableSixYdata, Predict_TableSevenYdata = get_former_Ydata(
    #         Preddata)
    #     all_original_Ydata = [Original_TableTwoYdata, Original_TableThreeYdata, Original_TableFiveYdata,
    #                           Original_TableSixYdata, Original_TableSevenYdata]
    #     all_predict_Ydata = [Predict_TableTwoYdata, Predict_TableThreeYdata, Predict_TableFiveYdata,
    #                          Predict_TableSixYdata, Predict_TableSevenYdata]
    #     # %%
    #     # 将Y取出来，然后取出缺失值过多的列
    #     # 获取可用的列
    #     Original_TableYdata, Predict_TableYdata = all_original_Ydata[i], all_predict_Ydata[i]
    #     final_cols = get_final_useablecols(Original_TableYdata, Predict_TableYdata)
    #     # final_cols
    #     # %%
    #     # 获取可以用的Y_data
    #     Original_YdataList, Pred_YdataList = all_ydata(final_cols, Original_TableYdata, Predict_TableYdata)
    #     # 合并数据，训练的时候一起
    #     # all_train_y_data = np.r_[Original_YdataList, Pred_YdataList]
    #     # %%
    #     # 挨个遍历，取出有用的Y_data
    #     # 传入起始的坐标 24,67,108,145,179
    #     # 传入起始的坐标新表 22,31,50,71,88
    #     colum_list = [22, 31, 50, 71, 88]
    #     the_column = colum_list[i]
    #     start = 0
    #     for index in final_cols:
    #         # 建模的Y_data
    #         Modeling_Y_data = Original_YdataList[start]
    #         # all_train_y_data全部用来训练
    #         # Modeling_Y_data = Original_YdataList[:int(Pred_YdataList.shape[0] / 2)+1, :][start]
    #         # 预测的Y_data，取一半测试，一半预测
    #         Pred_Y_data = Pred_YdataList[start]
    #         # Pred_Y_data = Pred_YdataList[int(Pred_YdataList.shape[0] / 2) :, :][start]
    #         start += 1
    #         # 处理建模的X_data,Y_data
    #         temp_OriginalXdata = np.column_stack((base_Xdata, Modeling_Y_data))
    #         # 删除缺失的行
    #         temp_OriginalXdata = Del_deletion_data(temp_OriginalXdata, 0)
    #         # X_data,Y_data 是第一个表处理好的x,y
    #         X_data = temp_OriginalXdata[:, :-1]
    #         Y_data = temp_OriginalXdata[:, -1]
    #         # 处理验证的X_data,Y_data
    #         # 合并数据
    #         temp_verifydata = np.column_stack((Verify_Xdata, Pred_Y_data))
    #         # 删除缺失的行
    #         temp_verifydata = Del_deletion_data(temp_verifydata, 0)
    #         # 获取预测数据的X_data，Y_data
    #         X_verify_data = temp_verifydata[:, :-1]
    #         Y_verify_data = temp_verifydata[:, -1]
    #         print(list(set(np.where(np.isnan(temp_OriginalXdata.astype(float)) == True)[0].tolist())),
    #               list(set(np.where(np.isnan(temp_verifydata.astype(float)) == True)[0].tolist())))
    #
    #         # 作图
    #         # for x_item in range(0,X_data.shape[1]):
    #         #     # write_image(X_left=X_data[:,x_item], X_right=X_verify_data[:,x_item], names=f"表{i}的{x_item}列")
    #         #     write_images(X_left=X_data[:,x_item], X_right=X_verify_data[:,x_item], names=f"images/表{i}的{x_item}列.png")
    #
    #         if 1 == 1:
    #             # 将建模Y_data分类(-1,0,1)，并且取出边界值
    #             # 保留原始的数据
    #             Y_data_original = copy.deepcopy(Y_data)
    #             Y_data, Y_data_boundsMin, Y_data_boundsMax = Deal_sorted_Ydata(data=Y_data)
    #             # 获取预测数据的Y_data Verify_Ydata = Pred_Y_data
    #             # 保留原始的数据
    #             Verify_Ydata_original = copy.deepcopy(Y_verify_data)
    #             Verify_Ydata = deal_verify_ydata(Y_verify_data, Y_data_boundsMin, Y_data_boundsMax)
    #             # 查看是否替换成功
    #             print([unique(Y_data), Y_data_boundsMin, Y_data_boundsMax], np.unique(Verify_Ydata))
    #             # 切分训练数据和测试数据
    #
    #             ## 30%测试数据，70%训练数据，stratify=y表示训练数据和测试数据具有相同的类别比例 修改为0.25
    #             # X_train, X_test, y_train, y_test = train_test_split(X_data, Y_data, test_size=0.25, random_state=4,stratify=Y_data)
    #             # 手动划分数据集
    #             X_train, X_test, y_train, y_test = X_data, X_verify_data[:int(X_verify_data.shape[0] / 2) + 1,
    #                                                        :], Y_data, Y_verify_data[
    #                                                                    :int(Y_verify_data.shape[0] / 2) + 1]
    #             # 验证集
    #             X_verify_data = X_verify_data[int(X_verify_data.shape[0] / 2) + 1:, :]
    #             Y_verify_data = Y_verify_data[int(Y_verify_data.shape[0] / 2) + 1:]
    #             from sklearn.metrics import accuracy_score
    #             # 开始训练模型
    #             from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
    #
    #             # 设置训练次数
    #             train_numbers = 300
    #             # 获取模型  随机森林
    #             Rfc = RandomForestClassifier()
    #             model_name = "Rfc"
    #             save_model(Rfc, model_name, X_train, X_test, y_train, y_test, train_numbers)
    #
    #             # 随便取一个作为训练集的结果
    #             current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
    #             Rfc_train_acc, Rfc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)
    #
    #             Gbc = GradientBoostingClassifier()
    #             model_name = "Gbc"
    #             save_model(Gbc, model_name, X_train, X_test, y_train, y_test, train_numbers)
    #             current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
    #             Gbc_train_acc, Gbc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)
    #             from sklearn.svm import SVC
    #
    #             Svc = SVC()
    #             model_name = "Svc"
    #             save_model(Svc, model_name, X_train, X_test, y_train, y_test, train_numbers)
    #             current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
    #             Svc_train_acc, Svc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)
    #
    #             Rfc_pred = get_pred_data("Rfc", X_verify_data)
    #             Gbc_pred = get_pred_data("Gbc", X_verify_data)
    #             Svc_pred = get_pred_data("Svc", X_verify_data)
    #
    #             verify_y1 = accuracy_score(Y_verify_data.astype(float), Rfc_pred)
    #             verify_y2 = accuracy_score(Y_verify_data.astype(float), Gbc_pred)
    #             verify_y3 = accuracy_score(Y_verify_data.astype(float), Svc_pred)
    #
    #             verify_dic = {}
    #             # 取分数最高的点保存
    #             max_scores = max(verify_y1, verify_y2, verify_y3)
    #
    #             # gen_images(iter_left_data=None, iter_right_verify=None, files_name="final_results")
    #             write_images(data_left=Y_data_original, data_right=Verify_Ydata_original,
    #                          names=f"final_results/{all_titles[the_column + index - 1]}.png", score=max_scores)
    #
    #             # 计算准确率
    #             print('验证集Rfc_acc准确率:', verify_y1)
    #             print('验证集Gbc_acc准确率:', verify_y2)
    #             print('验证集SVC准确率：', verify_y3)
    #
    #             #  标签 都加20
    #             write_to_excel(start_index=5, table=table, current_column=1,
    #                            content=["训练集个数", "随机森林（Rfc）准确率", "梯度提升树（Gbc）准确率",
    #                                     "支持向量机(Svc)准确率", "测试集个数", "随机森林（Rfc）准确率",
    #                                     "梯度提升树（Gbc）准确率", "支持向量机(Svc)准确率", "验证集个数",
    #                                     "随机森林（Rfc）准确率", "梯度提升树（Gbc）准确率", "支持向量机(Svc)准确率", ])
    #             # 将训练集写入到excel中
    #             write_to_excel(start_index=5, table=table, current_column=the_column + index,
    #                            content=[X_train.shape[0], Rfc_train_acc, Gbc_train_acc, Svc_train_acc, X_test.shape[0],
    #                                     Rfc_acc, Gbc_acc, Svc_acc,
    #                                     X_verify_data.shape[0], verify_y1, verify_y2, verify_y3])
    #
    #             data.save(filename)
